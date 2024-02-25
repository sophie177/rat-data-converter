'''
rat data converter for the Neuro Lab at WSU (commissioned by Damien Lybrand) 
- goal: convert file of delimited data to a sorted and readable version 
- current version: prototype
- current capabilities: sort data for one rat 
- future tasks: input checking, error handling for incomplete files, sort several rats' data for multiple days. 

'''
import openpyxl
#import pandas as pd

def sort_excel_data(input_file, output_file):
    # Load the Excel file
    workbook = openpyxl.load_workbook(input_file)# Assume the first sheet in the workbook
    sheet = workbook.active
    
    found_box = False  # for debugging
    found_right_lever = False # for debugging
    start_extracting = False
    box_num = 0 # in raw data, boxes start at 1 so this doubles as an error flag. 

    # Create a new workbook and sheet
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active

    left_lever_data = []
    left_totals = []
    right_lever_data = []
    right_totals = []
    
    # Iterate through each cell in the original sheet: First idenfity 'box' delimiter
    for row in sheet.iter_rows(values_only=True):
        
        for cell_value in row:
            if 'Box:' in str(cell_value):
                found_box = True
                box_num += 1  # if box is found, increment box num.
                new_sheet.append([cell_value]) # append value of box num to the new sheet
                new_sheet.append([box_num]) # incremented to keep track
            if found_box:
                if 'C:' in str(cell_value):
                   # new_sheet.append([cell_value]) 
                    start_extracting = True
                    continue                         
                if 'D:' in str(cell_value):
                    #new_sheet.append([cell_value]) 
                    found_right_lever = True
                    start_extracting = False
                    
                if start_extracting and isinstance(cell_value, (int, float)) and ':' not in str(cell_value):
                        left_lever_data.append(cell_value)
                                     
                if found_right_lever and isinstance(cell_value, (int, float)) and ':' not in str(cell_value):
                        right_lever_data.append(cell_value)
                        
    # Create Second Column: Running Totals
    running_total = 0
    for value in left_lever_data:
        running_total += value
        left_totals.append(round(running_total/1, 2)) #        left_totals.append(running_total)

        
    running_total = 0 # re-assign
    for value in right_lever_data:
        running_total += value 
        right_totals.append(round(running_total/1, 2))
    
    #running_total_divided = [value / 60 for value in left_totals] # adjust to three decimal places! 
    running_total_divided1 = [round(value / 100, 2) for value in left_totals]
    # Append the 'L:' column and 'Running Total' column to the new sheet
    new_sheet.append(["C:"] + ["C Totals"] + [" C seconds"]) # column headers
    for value, total, total_divided in zip(left_lever_data, left_totals, running_total_divided1):
        new_sheet.append([value, total, total_divided]) # was [ value, total, total_divided]
        
    running_total_divided2 = [round(value / 100, 2) for value in right_totals]

    new_sheet.append(["D:"] + ["D Totals"] + ["D seconds"])
    for value, total, total_divided in zip(right_lever_data, right_totals, running_total_divided2):
        new_sheet.append([value, total, total_divided])
        
    running_total_60 = [round(value / 60, 2) for value in running_total_divided1]
    new_sheet.append(["C:"] + ["C Totals"] + ["C Minutes"])

    for value, total, total_divided in zip(left_totals, running_total_divided1, running_total_60 ):
        new_sheet.append([value, total, total_divided]) 
        
    running_total_60_2 = [round(value/60, 2)for value in running_total_divided2]
    new_sheet.append(["D:"] + ["D Totals"] + ["D Minutes"])

    for value, total, total_divided in zip(right_totals, running_total_divided2, running_total_60_2):
        new_sheet.append([value, total, total_divided])
    

    
    # ~~~  Insert single-cell data lists for each lever so Damien can copy-paste into Matlab ~~~ 
    new_sheet.append(["C ~ Raw"])
    left_lever_cell = ', '.join(map(str, left_lever_data))
    new_sheet.append([left_lever_cell])
    
    new_sheet.append(["D ~ Raw"])
    right_lever_cell = ', '.join(map(str, right_lever_data))
    new_sheet.append([right_lever_cell])
    
    # single-cell lists of respective running totals 
    new_sheet.append(["C ~ Running Totals"])
    left_totals_cell = ', '.join(map(str, left_totals))
    new_sheet.append([left_totals_cell])
    
    new_sheet.append(["D ~ Running Totals"])
    right_totals_cell = ', '.join(map(str, right_totals))
    new_sheet.append([right_totals_cell])
    
    leftmin_list = [round(value / 60, 3) for value in left_totals]

    # single-sheet lists of respective running totals / 60
    new_sheet.append(["C ~ Minutes"])
    left_minutes_cell = ', '.join(str(value) for value in leftmin_list)
    new_sheet.append([left_minutes_cell])  # Append the comma-separated string as a single cell

    rightmin_list= [round(value / 60, 3) for value in right_totals]

    new_sheet.append(["D ~ Minutes"])
    right_minutes_cell = ', '.join(str(value) for value in rightmin_list)
    new_sheet.append([right_minutes_cell])  # Append the comma-separated string as a single cell



    # Save the new workbook to the output file
    new_workbook.save(output_file)


if __name__ == "__main__":
    # Provide the input and output file paths
    print(" Hello, Welcome to RatSort! ")
    print(" RatSort will require an input file path and return the path of the updated file.")

    input_file_path = input("paste or type input file path here: ")

    # input_file_path = "input_data.xlsx"
    output_file_path = "sorted_data.xlsx"

    # Call the function to sort the data
    sort_excel_data(input_file_path, output_file_path)

    print(f"Data has been sorted and saved to {output_file_path}")
    print("Check the folder of this program for your sorted Data. ")
